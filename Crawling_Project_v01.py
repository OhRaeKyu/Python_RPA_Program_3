# 웹 크롤링 프로그램 v01
# - 네이버 쇼핑, 블로그, 카페 웹 페이지 등에서 입력 받은 검색어에 대한 연관 검색어를 크롤링하는 프로그램
# 작성자 : 오래규
# 작성일 : 2021-02-22

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
import time
import os

LOAD_TIME = 1 # 드라이버 각 행동 간 로딩 시간

key = str(input("검색어를 입력하시오 : "))

wb = load_workbook("키워드 크롤링.xlsx")  # 준비된 엑셀 파일 로드
ws = wb.copy_worksheet(wb["양식"])  # 양식 시트 복사하여 새로운 시트 생성
ws.title = key

driver = webdriver.Chrome() # 크롬 브라우저 사용

#################################### 네이버 메인 페이지 작업 ####################################
driver.get("https://www.naver.com")

time.sleep(LOAD_TIME)

query = driver.find_element_by_name("query")
query.send_keys(key)

time.sleep(LOAD_TIME)

# 네이버 메인 자동완성 데이터 추출
main_xpath = "/html/body/div[2]/div[2]/div[1]/div/div[3]/div[3]/div/div/div[2]/div[1]/ul/li"  # 메인 페이지 자동완성 path
main_kwds = driver.find_elements_by_xpath(main_xpath)

row_m = 3
for main_kwd in main_kwds:
  main_data = main_kwd.get_attribute("data-keyword")
  ws.cell(column = 1, row = row_m).value = main_data  # 메인 페이지 자동완성 데이터 엑셀 입력
  row_m += 1


#################################### 네이버 쇼핑 페이지 작업 ####################################
driver.get("https://shopping.naver.com")

time.sleep(LOAD_TIME)

query = driver.find_element_by_name("query")
query.send_keys(key)
query.send_keys(Keys.RETURN)

time.sleep(LOAD_TIME)

# 네이버 쇼핑 연관검색어 데이터 추출
driver.find_element_by_css_selector(".relatedTags_btn_more__3Um4x").click() # 연관검색어 더보기 버튼 클릭

shop_xpath = "/html/body/div/div/div[2]/div[1]/div/ul/li" # 쇼핑 페이지 연관검색어 path
shop_kwds = driver.find_elements_by_xpath(shop_xpath)

row_s = 3
for shop_kwd in shop_kwds:
  shop_data = shop_kwd.find_element_by_tag_name("a").text
  ws.cell(column = 2, row = row_s).value = shop_data  # 쇼핑 페이지 연관검색어 데이터 엑셀 입력
  row_s += 1

# 네이버 쇼핑 상품 제목 추출
PG_cnt = 0
MAX_PG = 10 # [추출하고자 하는 페이지 수]

title_dict = dict()
title_list = list()

last_height = driver.execute_script("return document.body.scrollHeight")  # 스크롤이 끝까지 내려갔는지 확인하기 위한 변수
while PG_cnt < MAX_PG:
  driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")  # 스크롤 내리기
  time.sleep(LOAD_TIME) # 로딩 시간 대기
  new_height = driver.execute_script("return document.body.scrollHeight") # 스크롤이 끝까지 내려갔는지 확인하기 위한 변수

  if new_height == last_height:
    shop_titles = driver.find_elements_by_css_selector(".basicList_link__1MaTN")  # 상품 제목 class로 접근
    for shop_title in shop_titles:
      title_data = shop_title.get_attribute("title")
      title_splits = title_data.split(" ")  # 띄어쓰기 기준으로 상품 제목 분할
      for title_split in title_splits:
        title_dict[title_split] = title_dict.get(title_split, 0) + 1
    driver.find_element_by_css_selector(".pagination_next__1ITTf").click()
    PG_cnt += 1
  last_height = new_height

for k, v in title_dict.items():
  title_list.append((v, k))

title_list = sorted(title_list, reverse = True) # 내림차순 정렬

row_t = 3
for v, k in title_list:
  ws.cell(column = 5, row = row_t).value = k
  ws.cell(column = 6, row = row_t).value = v
  row_t += 1

#################################### 네이버 블로그 페이지 작업 ####################################
driver.get("https://section.blog.naver.com/")

time.sleep(LOAD_TIME)

query = driver.find_element_by_name("sectionBlogQuery")
query.send_keys(key)
query.send_keys(Keys.RETURN)

time.sleep(LOAD_TIME)

# 네이버 블로그 연관검색어 추출
blog_xpath = "/html/body/ui-view/div/main/div/aside/div/div[2]/div[2]/a"
blog_kwds = driver.find_elements_by_xpath(blog_xpath)

row_b = 3
for blog_kwd in blog_kwds:
  blog_data = blog_kwd.text
  ws.cell(column = 3, row = row_b).value = blog_data  # 블로그 페이지 연관검색어 데이터 엑셀 입력
  row_b += 1


#################################### 네이버 카페 페이지 작업 ####################################
driver.get("https://section.cafe.naver.com/")

time.sleep(LOAD_TIME)

query = driver.find_element_by_css_selector(".snb_search_text")
query.send_keys(key)
query.send_keys(Keys.RETURN)

time.sleep(LOAD_TIME)

# 네이버 카페 연관검색어 추출
driver.find_element_by_css_selector(".button_list_open_close").click() # 연관 검색어 더보기 버튼 클릭

cafe_xpath = "/html/body/div/div/div[2]/div/div[1]/div[1]/div/div/a"
cafe_kwds = driver.find_elements_by_xpath(cafe_xpath)

row_c = 3
for cafe_kwd in cafe_kwds:
  cafe_data = cafe_kwd.text
  ws.cell(column = 4, row = row_c).value = cafe_data  # 카페 페이지 연관검색어 데이터 엑셀 입력
  row_c += 1

wb.save("키워드 크롤링.xlsx")
driver.close()
os.system("pause")